#!/usr/bin/env python3
"""
Translation update script.

Reads the Excel files produced by Weblate and applies their content
to the JSON translation files used by the patient app.

Excel sheet columns (per Weblate export):
    A: location
    B: source (English text)
    C: target (translated text)
    D: id
    E: fuzzy
    F: context (dot-notation key, e.g. "cgu.backToHome")
    G: translator_comments
    H: developer_comments

Behavior:
    - Arabic translation.xlsx  -> updates src/assets/i18n/ar.json (deep merge)
    - Pashto translation.xlsx  -> creates src/assets/i18n/ps.json
    - Somali translation.xlsx  -> creates src/assets/i18n/so.json
    - Adds "ps" and "so" language labels to every i18n/*.json file
"""

import json
import os
from pathlib import Path

import openpyxl


HERE = Path(__file__).resolve().parent
I18N_DIR = HERE / "src" / "assets" / "i18n"

# Native labels for the newly supported languages.
NEW_LANGUAGE_LABELS = {
    "ps": "پښتو",
    "so": "Soomaali",
}

# (excel filename, target json filename, mode)
JOBS = [
    ("Arabic translation.xlsx", "ar.json", "merge"),
    ("Pashto translation.xlsx", "ps.json", "create"),
    ("Somali translation.xlsx", "so.json", "create"),
]


def read_excel_translations(xlsx_path: Path) -> dict:
    """Return a flat dict {dot.key: translation} from a Weblate export."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    ws = wb[wb.sheetnames[0]]

    flat = {}
    rows = ws.iter_rows(values_only=True)
    header = next(rows, None)
    if header is None:
        return flat

    # Column indices for "target" and "context".
    target_idx = 2
    context_idx = 5

    for row in rows:
        if not row:
            continue
        key = row[context_idx]
        value = row[target_idx]
        if not key or value is None:
            continue
        key = str(key).strip()
        value = str(value).strip()
        if not key or not value:
            continue
        flat[key] = value
    return flat


def nest(flat: dict) -> dict:
    """Convert {"a.b.c": "x"} into {"a": {"b": {"c": "x"}}}."""
    nested = {}
    for dotted_key, value in flat.items():
        parts = dotted_key.split(".")
        cursor = nested
        for part in parts[:-1]:
            existing = cursor.get(part)
            if not isinstance(existing, dict):
                existing = {}
                cursor[part] = existing
            cursor = existing
        cursor[parts[-1]] = value
    return nested


def deep_merge(base: dict, overlay: dict) -> dict:
    """Recursively merge overlay into base. Overlay values win on conflict."""
    out = dict(base)
    for key, value in overlay.items():
        if (
            key in out
            and isinstance(out[key], dict)
            and isinstance(value, dict)
        ):
            out[key] = deep_merge(out[key], value)
        else:
            out[key] = value
    return out


def load_json(path: Path) -> dict:
    if not path.exists():
        return {}
    with path.open("r", encoding="utf-8") as f:
        return json.load(f)


def save_json(path: Path, data: dict) -> None:
    with path.open("w", encoding="utf-8") as f:
        json.dump(data, f, indent=4, ensure_ascii=False)
        f.write("\n")


def process_job(xlsx_name: str, json_name: str, mode: str) -> None:
    xlsx_path = HERE / xlsx_name
    json_path = I18N_DIR / json_name

    if not xlsx_path.exists():
        print(f"  [SKIP] Missing {xlsx_path.name}")
        return

    flat = read_excel_translations(xlsx_path)
    nested = nest(flat)

    if mode == "merge" and json_path.exists():
        existing = load_json(json_path)
        merged = deep_merge(existing, nested)
        save_json(json_path, merged)
        print(
            f"  [UPDATED] {json_name} "
            f"({len(flat)} entries from Excel applied on top of existing keys)"
        )
    else:
        save_json(json_path, nested)
        action = "CREATED" if not json_path.exists() else "REPLACED"
        # path.exists() was true after save; report based on mode instead.
        action = "CREATED" if mode == "create" else "REPLACED"
        print(f"  [{action}] {json_name} ({len(flat)} entries)")


def patch_language_labels() -> None:
    """Ensure every i18n/*.json knows the native label for new languages."""
    for json_file in sorted(I18N_DIR.glob("*.json")):
        data = load_json(json_file)
        changed = False
        for code, label in NEW_LANGUAGE_LABELS.items():
            if data.get(code) != label:
                data[code] = label
                changed = True
        if changed:
            save_json(json_file, data)
            print(f"  [LABELS] {json_file.name} updated")


def main() -> None:
    print("== Processing Excel translation files ==")
    for xlsx_name, json_name, mode in JOBS:
        process_job(xlsx_name, json_name, mode)

    print("\n== Patching native language labels in every i18n file ==")
    patch_language_labels()

    print("\nDone.")


if __name__ == "__main__":
    main()
